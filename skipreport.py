import openpyxl
import warnings
import jinja2

reportxls = input("report xls: ").strip('"') #or 'skiplist2.xlsx'
if reportxls == 'd' : reportxls = 'skiplist.xlsx'
if not reportxls:
    print('no report to work on')
    exit()


# avoid useless warning: UserWarning: Discarded range with reserved name  warnings.warn("Discarded range with reserved name")
warnings.simplefilter("ignore")
wb = openpyxl.load_workbook(reportxls, read_only=True)
# put a url from CAPS in Sheet8.A1 if you want links in report
capsurl = wb.get_sheet_by_name('Sheet8').cell(row=1, column=1).value
if capsurl:
    proj = capsurl.split('/')[7]
    cont = capsurl.split('/')[9]
sheet = wb.get_sheet_by_name('All Skips')
docset = sheet.cell(row=2, column=4).value

skips = []; badtitles = {}; badtables = {}; badskips = {}
for r in range(2,sheet.max_row):
    if sheet.cell(row=r, column=7).value != 'en-US': continue
    Name  = sheet.cell(row=r, column=2).value
    Id    = sheet.cell(row=r, column=3).value
    AInfo = sheet.cell(row=r, column=9).value
    ai = AInfo.split(';')
    for bad in ai:
        bs = bad.strip()
        # print(bs)
        if 'Skip: Topic links to topic' in bs:
            badtitles[bs[27:63]] = 'bogus'
        elif 'contains complicated table' in bs:
            badtables[Id] = Name
        elif  'Warning:' in bs:
            continue
        elif  'is already MarkDown' in bs:
            continue
        else:
            badskips[Id] = bs
            print('unhandled skip:  ' + bs)

warnings.simplefilter("default") # done with openpyxl warnings

def topicurl(id):
    # CAPS url from the docset  e.g. "https://caps-web-prod.azurewebsites.net/#/organizations/e6f6a65cf14f462597b64ac058dbe1d0/projects/cb875318-aadb-47d1-98e5-fa163edd48a7/containers/8ab24a20-ae68-4357-8e3c-c92ca266f464/tocpaths/6e3c7844-a207-dbc6-39da-0ed37546c910/locales/en-US";
    proj = capsurl.split('/')[7]
    cont = capsurl.split('/')[9]
    capsprefix = 'https://caps-web-prod.azurewebsites.net/#/organizations/e6f6a65cf14f462597b64ac058dbe1d0'
    url = capsprefix + "/projects/" + proj + "/containers/" + cont + "/articles/" + id + "/locales/en-US"
    return url

templateLoader = jinja2.FileSystemLoader(searchpath=".")
templateEnv = jinja2.Environment(loader=templateLoader)
if capsurl:
    proj = capsurl.split('/')[7]
    cont = capsurl.split('/')[9]
    capsprefix = 'https://caps-web-prod.azurewebsites.net/#/organizations/e6f6a65cf14f462597b64ac058dbe1d0'
    bigCAPSurl = capsprefix + "/projects/" + proj + "/containers/" + cont + "/articles/"
    # + id + "/locales/en-US"
    TEMPLATE_FILE = "capslinks.html"
else:
    bigCAPSurl = 'bogus'
    TEMPLATE_FILE = "capsplain.html"

template = templateEnv.get_template(TEMPLATE_FILE)
             

templateVars = { "title" : "CAPS Report",
                 "url" : bigCAPSurl,
                 "docset" : docset,
                 "badtitles" : badtitles,
                 "badtables" : badtables,
                 "badskips" : badskips,
               }

outputText = template.render(templateVars)
fout = open('report.html', 'wt')
print(outputText, file=fout)
fout.close()


# Skip: Topic links to topic c8836345-16bb-4dcc-8d2b-2b9b687456a3 with illegal name. Please remove / or \ in the name.
# Skip: contains complicated table

# Dupe titles.
# Names with bad characters
# Nested tables (table inside table)
# Code in tables
# Tables with colspan and/or rowspan greater than 1

